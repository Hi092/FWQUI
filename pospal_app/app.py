#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
【一杯时间】销售日报 Web 应用（简化版）
收款方式：现金、银豹付（微信），总额=营业实收
"""

import os, re, logging, datetime
from io import BytesIO
from flask import Flask, render_template, request, jsonify, send_file
import requests
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

BASE_URL = "https://beta.pospal.cn"
LOGIN_URL = f"{BASE_URL}/Account/SignIn"
PRODUCT_API = f"{BASE_URL}/ReportV2/LoadProductSaleByPage"
PAYMENT_API = f"{BASE_URL}/ReportV2/LoadStorePaymentSummary"

ACCOUNT = os.environ.get("POSPAL_ACCOUNT", "15765588244")
EMP_ID = os.environ.get("POSPAL_EMP_ID", "S001")
PASSWORD = os.environ.get("POSPAL_PASSWORD", "1234")
DEFAULT_USER_ID = "4899673"

SESSION_CACHE = {"session": None, "user_id": None, "expire": None}
REPORT_CACHE = {}


def safe_float(text):
    if text is None:
        return 0.0
    try:
        return float(str(text).strip().replace(",", ""))
    except:
        return 0.0


def safe_int(text):
    return int(round(safe_float(text)))


def build_session():
    s = requests.Session()
    retry = Retry(total=3, backoff_factor=1, status_forcelist=[500, 502, 503, 504])
    s.mount("http://", HTTPAdapter(max_retries=retry))
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.headers.update({"User-Agent": "Mozilla/5.0"})
    return s


def create_session():
    now = datetime.datetime.now()
    if SESSION_CACHE["session"] and SESSION_CACHE["expire"] and SESSION_CACHE["expire"] > now:
        return SESSION_CACHE["session"], SESSION_CACHE["user_id"], None

    logger.info("重新登录银豹")
    session = build_session()
    try:
        resp = session.get(LOGIN_URL, timeout=15, allow_redirects=True)
        resp.raise_for_status()
    except Exception as e:
        return None, None, f"连接银豹失败: {e}"

    store = ""
    m = re.search(r"store=([^&]+)", resp.url)
    if m:
        store = m.group(1)

    if "EmployeeSignin" in resp.url:
        login_url = f"{BASE_URL}/account/EmployeeSignin?store={store}"
        login_data = {"userName": EMP_ID, "password": PASSWORD, "screenSize": "1920*1080"}
    else:
        login_url = f"{BASE_URL}/account/SignIn?noLog="
        login_data = {"userName": f"{ACCOUNT}:{EMP_ID}", "password": PASSWORD, "employeeSignin": "true", "screenSize": "1920*1080"}

    try:
        resp = session.post(login_url, data=login_data, timeout=15, allow_redirects=True,
                            headers={"Content-Type": "application/x-www-form-urlencoded", "Referer": resp.url, "X-Requested-With": "XMLHttpRequest"})
        result = resp.json()
        if not result.get("successed"):
            return None, None, f"登录失败: {result.get('msg')}"
    except Exception as e:
        return None, None, f"登录异常: {e}"

    user_id = DEFAULT_USER_ID
    try:
        test = session.get(f"{BASE_URL}/ReportV2/ProductSale", timeout=15)
        m = re.search(r"userIds.*?(\d{6,})", test.text)
        if m:
            user_id = m.group(1)
    except:
        pass

    SESSION_CACHE.update({"session": session, "user_id": user_id, "expire": now + datetime.timedelta(minutes=30)})
    logger.info("银豹登录成功")
    return session, user_id, None


def get_product_sales(session, user_id, target_date):
    begin = target_date.strftime("%Y.%m.%d 00:00:00")
    end = target_date.strftime("%Y.%m.%d 23:59:59")
    body = (f"groupByArtNo=false&keyword=&userIds%5B%5D={user_id}&isSellWell=1"
            f"&beginDateTime={requests.utils.quote(begin)}&endDateTime={requests.utils.quote(end)}"
            f"&pageIndex=1&pageSize=500")
    headers = {"Content-Type": "application/x-www-form-urlencoded", "Referer": f"{BASE_URL}/ReportV2/ProductSale", "X-Requested-With": "XMLHttpRequest"}

    resp = session.post(PRODUCT_API, data=body, headers=headers, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    if not data.get("successed"):
        return []

    soup = BeautifulSoup(data.get("contentView", ""), "lxml")
    products = []
    for tr in soup.select("tr[data]"):
        tds = tr.find_all("td")
        if len(tds) < 14:
            continue
        try:
            name = tds[3].get_text(strip=True)
            unit = tds[7].get_text(strip=True)
            qty = safe_int(tds[12].get_text(strip=True))
            total = safe_float(tds[13].get_text(strip=True))
            if not name or qty <= 0:
                continue
            unit_price = round(total / qty, 2)
            products.append({"name": name, "unit": unit, "qty": qty, "unit_price": unit_price, "total": total})
        except Exception as e:
            logger.warning(f"商品解析失败: {e}")
    return products


def get_payment_summary(session, user_id, target_date):
    date_str = target_date.strftime("%Y-%m-%d")
    body = f"beginDateTime={date_str}&endDateTime={date_str}&userIds%5B%5D={user_id}"
    headers = {"Content-Type": "application/x-www-form-urlencoded", "Referer": f"{BASE_URL}/ReportV2/StorePaymentSummary", "X-Requested-With": "XMLHttpRequest"}

    resp = session.post(PAYMENT_API, data=body, headers=headers, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    if not data.get("successed"):
        return {"cash": 0, "wechat": 0, "total": 0}

    rows = data.get("json", {}).get("list", [])
    if not rows:
        return {"cash": 0, "wechat": 0, "total": 0}

    row = rows[0]
    return {
        "cash": safe_float(row.get("现金支付", {}).get("amount", 0)),
        "wechat": safe_float(row.get("银豹付支付", {}).get("amount", 0)),
        "member": safe_float(row.get("储值卡支付", {}).get("amount", 0)),
        "total": safe_float(row.get("营业实收", 0))
    }


# ========================= 商品收款方式查询 =========================
def get_cash_products(session, user_id, target_date):
    """从销售单据API获取每个商品的现金/微信金额（按小计分配）"""
    begin = target_date.strftime("%Y.%m.%d 00:00:00")
    end = target_date.strftime("%Y.%m.%d 23:59:59")
    body = (f"userIds%5B%5D={user_id}&beginTime={begin}&endTime={end}"
            f"&reversed=0&onlyCustomer=false&onlyWholesale=false&onlyReturn=false"
            f"&cashierUid=&guiderUid=&paymethod=&paymethodNames=null&sn="
            f"&tableUids%5B%5D=&pageIndex=1&pageSize=200&orderColumn=&asc=false"
            f"&cashCouponCode=&orderSource=&verificationSource=")

    headers = {"Content-Type": "application/x-www-form-urlencoded", "X-Requested-With": "XMLHttpRequest"}

    try:
        resp = session.post(f"{BASE_URL}/Report/LoadTicketsByPage", data=body, timeout=30, headers=headers)
        data = resp.json()
        if not data.get("successed"):
            return {}
    except:
        return {}

    soup = BeautifulSoup(data.get("contentView", ""), "lxml")
    rows = soup.find_all("tr")

    product_payments = {}
    current_products = []
    current_cash = 0
    current_wechat = 0

    for row in rows:
        classes = row.get("class", [])

        if row.find("a", class_="btn_showItems"):
            if current_products:
                ticket_total = sum(sub for _, sub in current_products)
                for name, subtotal in current_products:
                    if name not in product_payments:
                        product_payments[name] = {"cash": 0, "wechat": 0}
                    if ticket_total > 0:
                        ratio = subtotal / ticket_total
                        product_payments[name]["cash"] += current_cash * ratio
                        product_payments[name]["wechat"] += current_wechat * ratio

            current_products = []
            current_cash = 0
            current_wechat = 0

        elif "ticketItemRow" in classes:
            tds = row.find_all("td")
            text = row.get_text(strip=True)

            if "支付方式：" in text:
                match = re.search(r"支付方式[：:]\s*(\S+)\s+([\d.]+)", text)
                if match:
                    amount = float(match.group(2))
                    if "现金" in match.group(1):
                        current_cash += amount
                    else:
                        current_wechat += amount

            elif len(tds) == 8:
                product_text = tds[1].get_text(strip=True)
                if "(" in product_text and ")" in product_text:
                    name = product_text.split("(")[0].strip()
                    try:
                        subtotal = float(tds[4].get_text(strip=True))
                    except:
                        subtotal = 0
                    current_products.append((name, subtotal))

    if current_products:
        ticket_total = sum(sub for _, sub in current_products)
        for name, subtotal in current_products:
            if name not in product_payments:
                product_payments[name] = {"cash": 0, "wechat": 0}
            if ticket_total > 0:
                ratio = subtotal / ticket_total
                product_payments[name]["cash"] += current_cash * ratio
                product_payments[name]["wechat"] += current_wechat * ratio

    return product_payments


def generate_excel(products, target_date, payment):
    month, day = target_date.month, target_date.day
    filename = f"【一杯时间】销售日报表{month}月{day}日.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "销售日报"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"【一杯时间】销售日报表{month}月{day}日"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    for col, h in enumerate(["序号", "日期", "销售商品", "单位", "数量", "单价", "金额", "收款方式"], 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = border

    for i, p in enumerate(products, 1):
        row = i + 2
        for col, val in enumerate([i, target_date.strftime("%Y-%m-%d"), p["name"], p["unit"], p["qty"], p["unit_price"], p["total"], p.get("payment", "")], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = center

    r = len(products) + 4
    ws.cell(row=r, column=6, value="微信：").font = Font(bold=True)
    ws.cell(row=r, column=7, value=int(payment["wechat"]))
    ws.cell(row=r + 1, column=6, value="现金：").font = Font(bold=True)
    ws.cell(row=r + 1, column=7, value=int(payment["cash"]))
    ws.cell(row=r + 2, column=6, value="合计：").font = Font(bold=True)
    ws.cell(row=r + 2, column=7, value=int(payment["total"]))

    for col, w in {1: 8, 2: 16, 3: 36, 4: 10, 5: 8, 6: 12, 7: 12, 8: 14}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, filename


@app.route("/")
def index():
    return render_template("index.html", today=datetime.date.today(), yesterday=datetime.date.today() - datetime.timedelta(days=1))


@app.route("/api/report")
def api_report():
    date_str = request.args.get("date")
    if not date_str:
        target_date = datetime.date.today()
    else:
        try:
            target_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except:
            return jsonify({"error": "日期格式错误"}), 400

    force = request.args.get("refresh") == "1"
    cache_key = str(target_date)

    if not force and cache_key in REPORT_CACHE:
        cached = REPORT_CACHE[cache_key]
        if datetime.datetime.now() < cached["expire"]:
            logger.info(f"使用缓存 {cache_key}")
            return jsonify(cached["data"])

    session, user_id, err = create_session()
    if err:
        return jsonify({"error": err}), 500

    try:
        products = get_product_sales(session, user_id, target_date)
        payment = get_payment_summary(session, user_id, target_date)
    except Exception as e:
        logger.exception(e)
        return jsonify({"error": str(e)}), 500

    if not products:
        return jsonify({"error": "该日期无销售数据", "date": str(target_date), "products": [], "payment": payment})

    total_sales = sum(p["total"] for p in products)
    total_qty = sum(p["qty"] for p in products)

    member = payment.pop("member", 0)
    expected_all = payment["total"] + member

    # 获取每个商品的收款方式
    cash_product_names = get_cash_products(session, user_id, target_date)
    for p in products:
        name = p["name"]
        if cash_product_names and name in cash_product_names:
            cp = cash_product_names[name]
            cash_amt = int(cp["cash"])
            wechat_amt = int(cp["wechat"])
            if cash_amt > 0 and wechat_amt > 0:
                p["payment"] = f"微信/现金{cash_amt}"
            elif cash_amt > 0:
                p["payment"] = f"现金{cash_amt}"
            else:
                p["payment"] = "微信"
        else:
            p["payment"] = "微信"

    result = {
        "date": str(target_date),
        "products": products,
        "payment": payment,
        "summary": {
            "total_sales": total_sales,
            "total_qty": total_qty,
            "item_count": len(products),
            "expected": payment["total"],
            "verified": abs(total_sales - expected_all) <= 1
        }
    }

    REPORT_CACHE[cache_key] = {"data": result, "expire": datetime.datetime.now() + datetime.timedelta(minutes=5)}
    return jsonify(result)


@app.route("/api/download")
def api_download():
    date_str = request.args.get("date")
    if not date_str:
        target_date = datetime.date.today()
    else:
        try:
            target_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except:
            return jsonify({"error": "日期格式错误"}), 400

    # 直接获取最新数据，不用缓存
    session, user_id, err = create_session()
    if err:
        return jsonify({"error": err}), 500

    try:
        products = get_product_sales(session, user_id, target_date)
        payment = get_payment_summary(session, user_id, target_date)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if not products:
        return jsonify({"error": "该日期无销售数据"}), 400

    # 去掉 member 字段，标注收款方式
    payment.pop("member", 0)
    cash_product_names = get_cash_products(session, user_id, target_date)
    for p in products:
        name = p["name"]
        if cash_product_names and name in cash_product_names:
            cp = cash_product_names[name]
            cash_amt = int(cp["cash"])
            wechat_amt = int(cp["wechat"])
            if cash_amt > 0 and wechat_amt > 0:
                p["payment"] = f"微信/现金{cash_amt}"
            elif cash_amt > 0:
                p["payment"] = f"现金{cash_amt}"
            else:
                p["payment"] = "微信"
        else:
            p["payment"] = "微信"

    buf, filename = generate_excel(products, target_date, payment)
    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    logger.info("=" * 40)
    logger.info("【一杯时间】销售日报系统 v2")
    logger.info("http://0.0.0.0:8080")
    logger.info("=" * 40)
    app.run(host="0.0.0.0", port=8080, debug=False)
